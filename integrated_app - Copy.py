import os
import pathlib
import io
import base64
import numpy as np
import dash
from openpyxl import load_workbook
import dash_core_components as dcc
import dash_html_components as html
from dash.dependencies import Input, Output, State
import dash_table 
import plotly.graph_objs as go
import dash_daq as daq
import statsmodels.formula.api as sm
import scipy.special as ssp
from urllib.parse import quote as urlquote
from flask import Flask, send_from_directory
from plotly.subplots import make_subplots
import dash_bootstrap_components as dbc
import time
import pandas as pd

from functions_used import summary_stats 
from functions_used import contribution
# imports for integrating
import tab_1
import tab_2
import post_model_tab
import upload_data_tab

server = Flask(__name__)

app = dash.Dash(
    __name__,server=server,
    meta_tags=[{"name": "viewport", "content": "width=device-width, initial-scale=1"}],
    external_stylesheets= [dbc.themes.GRID]
)

server = app.server


j=2

APP_PATH = str(pathlib.Path("__file__").parent.resolve())
print("APP PATH is ")
print(APP_PATH)
UPLOAD_DIRECTORY = os.path.join(APP_PATH, os.path.join("data"))


app.config["suppress_callback_exceptions"] = True

@server.route("/download/<path:path>")
def download(path):
    """Serve a file from the upload directory."""
    return send_from_directory(UPLOAD_DIRECTORY, path, as_attachment=True)



df = pd.read_csv(os.path.join(APP_PATH, os.path.join("data", "spc_data.csv")))

col_names=['Variable','Coeff.','logit avg.','logit min.','y pred. logit','y pred. min','y pred act.','y pred. act. min','% diff','Normalized Contri.']
params = list(df)
print(params)


def build_banner():
    return html.Div(
        id="banner",
        className="banner",
        children=[
            html.Div(
                id="banner-text",
                children=[
                    html.H5("SWIFT MMM Dash"),
                    
                ],
            ),
            html.Div(
                id="banner-logo",
                children=[
                    html.Button(
                        id="learn-more-button", children="LEARN MORE", n_clicks=0
                    ),
                    html.Div(style={'display':'none'},children=[        
                    html.Div(id='file_name', style={'display': 'none'}),
                    dash_table.DataTable(id='dfl_table',
                                     #columns=([{'id': p, 'name': p} for p in col_names]),
                                     #data=generate_table_data(),                                   
                                     editable=False,                                     
                                     style_cell={'backgroundColor': 'rgb(8, 8, 8)','color': 'white','text-align':'center'},
                                     #style={'display':'none',}
                                     ),
                    dash_table.DataTable(id='dfa_table',
                                     #columns=([{'id': p, 'name': p} for p in col_names]),
                                     #data=generate_table_data(),                                   
                                     editable=False,                                     
                                     style_cell={'backgroundColor': 'rgb(8, 8, 8)','color': 'white','text-align':'center'},
                                     #style={'display':'none',}
                                     ),
                    dash_table.DataTable(id='summary_table',
                                     #columns=([{'id': p, 'name': p} for p in col_names]),
                                     #data=generate_table_data(),                                   
                                     editable=False,                                     
                                     style_cell={'backgroundColor': 'rgb(8, 8, 8)','color': 'white','text-align':'center'},
                                     #style={'display':'none',}
                                     ),
                    dash_table.DataTable(id='params_table',
                                     #columns=([{'id': p, 'name': p} for p in col_names]),
                                     #data=generate_table_data(),                                   
                                     editable=False,                                     
                                     style_cell={'backgroundColor': 'rgb(8, 8, 8)','color': 'white','text-align':'center'},
                                     #style={'display':'none',}
                                     ),
                    dash_table.DataTable(id='contri_table',
                                     columns=([{'id': p, 'name': p} for p in col_names]),
                                     data=generate_table_data(),                                   
                                     editable=False,                                     
                                     style_cell={'backgroundColor': 'rgb(8, 8, 8)','color': 'white','text-align':'center'},
                                     #style={'display':'none',}
                                     ),
                    dash_table.DataTable(id='temp_contri_table',
                                     columns=([{'id': p, 'name': p} for p in col_names]),
                                     data=generate_table_data(),                                   
                                     editable=False,                                     
                                     style_cell={'backgroundColor': 'rgb(8, 8, 8)','color': 'white','text-align':'center'},
                                     #style={'display':'none',}
                                     ),
                        
            dash_table.DataTable(id='avp',
                                     columns=([{'id': p, 'name': p} for p in col_names]),
                                     data=generate_table_data(),                                   
                                     editable=False,                                     
                                     style_cell={'backgroundColor': 'rgb(8, 8, 8)','color': 'white','text-align':'center'},
                                     #style={'display':'none',}
                                     ),
            dash_table.DataTable(id='rsqr',
                                     columns=([{'id': p, 'name': p} for p in col_names]),
                                     data=generate_table_data(),                                   
                                     editable=False,                                     
                                     style_cell={'backgroundColor': 'rgb(8, 8, 8)','color': 'white','text-align':'center'},
                                     #style={'display':'none',}
                                     ),  
                    
                    dcc.Input(id="kpi_name",style={'display':'none'}),                     
                    ],),
                    
                    
                    html.Img(id="logo", src=app.get_asset_url("PM_logo.png")),
                ],
            ),
        ],
    )


def build_tabs():
    return html.Div(
        id="tabs",
        className="tabs",
        children=[
            dcc.Tabs(
                id="app-tabs",
                value="upload_data",
                className="custom-tabs",
                children=[
#                    dcc.Tab(id="Specs-tab",label="Model Selection",value="tab1",className="custom-tab",
#                            selected_className="custom-tab--selected",),
                    dcc.Tab(id="upload_data",label="Upload Data",value="upload_data",className="custom-tab",
                            selected_className="custom-tab--selected"),
                    dcc.Tab(id='Specs-tab',label='Trend Analysis', value='tab-2-example',className="custom-tab",
                            selected_className="custom-tab--selected",),         
                    
                    dcc.Tab(id='Tab-1',label='Hypothesis Builder', value='tab-1-example',className="custom-tab",selected_className="custom-tab--selected",),
                    dcc.Tab(id="Control-chart-tab",label="Post Model Dashboard",value="tab2",className="custom-tab",selected_className="custom-tab--selected",),       
                ],
            )
        ],
    )

# =============================================================================
def generate_modal():
    return html.Div(
        id="markdown",
        className="modal",
        children=(
            html.Div(
                id="markdown-container",
                className="markdown-container",
                children=[
                    html.Div(
                        className="close-container",
                        children=html.Button(
                            "Close",
                            id="markdown_close",
                            n_clicks=0,
                            className="closeButton",
                        ),
                    ),
                    html.Div(
                        className="markdown-text",
                        children=dcc.Markdown(
                            children=(
                                """
                        ###### What is this app about?

                        This is a dashboard for building Marketing Mix Models using python.

                        ###### What does this app shows

                       Coming Soon......

                    """
                            )
                        ),
                    ),
                ],
            )
        ),
    )


def generate_section_banner(title):
    return html.Div(className="section-banner", children=title)



                                    

def generate_table_data():
    
    #pd.read_excel(os.path.join(APP_PATH, os.path.join("data", "")))
    
    return [dict({col: j for col in col_names})for i in range(1,2 )]
    

# Build header






app.layout = html.Div(
    id="big-app-container",
    children=[
        build_banner(),
        html.Div(
            id="app-container",
            children=[
                build_tabs(),
                # Main app
                html.Div(id="app-content"),
            ],
        ),
        
        
        generate_modal(),
    ],
)

#=======================callbacks for switching tabs==================
@app.callback(
    [Output("app-content", "children")],
    [Input("app-tabs", "value")],
    [State('file_name','children'),State('summary_table','data'),State('dfl_table','data'),
     State('dfa_table','data'),State('params_table','columns'),
     State('contri_table','data'),State('contri_table','columns')]
    
)

def render_tab_content(tab_switch,file_name,summary_dict,dfl_dict,dfa_dict,params_col,contri_data,contri_cols):
    if tab_switch == "upload_data":
        return (
        html.Div(
            id="status-container",
            children=[
                
                html.Div(
                    id="graphs-container",
                    children=[upload_data_tab.build_top_panel(contri_data,contri_cols)],
                ),
            ],
        ),

    )
    if tab_switch == "tab2":
        contri_df=pd.read_excel(os.path.join(APP_PATH, os.path.join("data", file_name['props']['children']+'_contribution.xlsx')))
        contri_data=contri_df.to_dict("records")
        contri_cols=([{'id': p, 'name': p} for p in contri_df.columns])
        return (
        html.Div(
            id="status-container",
            children=[
                
                html.Div(
                    id="graphs-container",
                    children=[post_model_tab.build_top_panel(contri_data,contri_cols)],
                ),
            ],
        ),

    )
    if tab_switch == 'tab-1-example':
        return (
        html.Div(
            id="status-container",
            children=[
                
                html.Div(
                    id="graphs-container",
                    children=[tab_1.tab1_layout(file_name,summary_dict,dfl_dict,dfa_dict,params_col)],
                ),
            ],
        ),

    )
    
    elif tab_switch == 'tab-2-example':
        return (
        html.Div(
            id="status-container",
            children=[
                
                html.Div(
                    id="graphs-container",
                    children=[tab_2.tab2_layout(file_name,summary_dict,dfl_dict,dfa_dict)],
                ),
            ],
        ),

    )      
    



# ======= Callbacks for modal popup =======
@app.callback(
    Output("markdown", "style"),
    [Input("learn-more-button", "n_clicks"), Input("markdown_close", "n_clicks")],
)
def update_click_output(button_click, close_click):
    ctx = dash.callback_context

    if ctx.triggered:
        prop_id = ctx.triggered[0]["prop_id"].split(".")[0]
        if prop_id == "learn-more-button":
            return {"display": "block"}

    return {"display": "none"}

#====================Callback for loading effect==============================
@app.callback(Output("loading-output-1", "children"), [Input("upload-data", "contents")])
def input_triggers_spinner(value):
    time.sleep(10)
    return 

# ====== Callbacks to upload data file via upload button =====
@app.callback(
    [Output("file_name", "children"),
     Output('summary_table','data'),Output('summary_table','columns'),
     Output('dfl_table','data'),Output('dfl_table','columns'),
     Output('dfa_table','data'),Output('dfa_table','columns'),Output('params_table','columns'),
     Output('kpi_name','value'),
     Output('upload_response_div','children')
     ],
    [Input("upload-data", "filename"), Input("upload-data", "contents")],
)
def upload_file_and_create_contri_data_table(uploaded_filenames,uploaded_file_contents):
    """Save uploaded files and regenerate the file list."""
    #print("Inside Call-back")
    
    if uploaded_filenames is None:
        print("No file uploaded")
    else:
        #print("Inside 1")
        for name,data in zip(uploaded_filenames, uploaded_file_contents):
            save_file(name, data)
            file_name=name
            #print(type(file_name))
        
        files = uploaded_files()
        if len(files) == 0:
            return html.Li("No files yet!"),file_name
        else:
            print("line 614")
            summary_dict,summary_cols, dfl_dict,dfl_cols, dfa_dict,dfa_cols,params_col,kpi_name=read_file_to_update_hiden_divs(file_name)
            generate_table_data_new(file_name,0)
            #contri_table_data_df=pd.DataFrame(data=contri_table_data,columns=contri_table_cols)
            
            #contri_table_data_df.to_excel(os.path.join(APP_PATH, os.path.join("data", file_name+"_contribution_table.xlsx")))
            
            return [html.Li(file_name),summary_dict,summary_cols, dfl_dict,dfl_cols, dfa_dict,dfa_cols,params_col,kpi_name,"Uploaded Sucessfully!"]

    return [html.Li("No files yet!"),file_name]

def read_file_to_update_hiden_divs(file_name):
    df = pd.read_excel(os.path.join(APP_PATH, os.path.join("data", file_name)),sheet_name="Data_Raw")
    inp_meta_data_df=pd.read_excel(os.path.join(APP_PATH, os.path.join("data", file_name)),sheet_name="KPI")
    time_grain=inp_meta_data_df['Time_Grain'][0]
    kpi_name=inp_meta_data_df['kpi_name'][0]
    

    df_params=pd.read_excel(os.path.join(APP_PATH, os.path.join("data", file_name)),sheet_name="Params")
    df_temp=df_params.filter(regex="Params_*",axis=1)
    df_params=df_params[df_params.columns[~df_params.columns.isin(df_temp.columns)]]
    df_temp=df_params.filter(regex="P_values_*",axis=1)
    df_params=df_params[df_params.columns[~df_params.columns.isin(df_temp.columns)]]
    df_params=df_params[df_params.columns[~df_params.columns.isin(['index','R_Square','Adj_R_Square'])]]
    variable_transformations=list(df_params.iloc[0]) 
    variable_transformations.insert(0,kpi_name)
    summary, dfl, dfa = summary_stats(df,time_grain)
    
    summary_dict=summary.to_dict("records")
    summary_cols=([{'id': p, 'name': p} for p in summary.columns])
    
    dfl_dict=dfl.to_dict("records")
    dfl_cols=([{'id': p, 'name': p} for p in dfl.columns])
    
    dfa_dict=dfa.to_dict("records")
    dfa_cols=([{'id': p, 'name': p} for p in dfa.columns])
    
    params_col=([{'id': p, 'name': p} for p in variable_transformations])
    return summary_dict,summary_cols, dfl_dict,dfl_cols, dfa_dict,dfa_cols,params_col,kpi_name

#   ------------------------------------Saving file to server directory inside data folder--------------------------
def save_file(name, content):
    """Decode and store a file uploaded with Plotly Dash."""
    
    print(name)
    #data = pd.read_excel(name)
    #df = pd.read_csv(os.path.join(APP_PATH, os.path.join("data", "spc_data.csv")))
    data=content.encode("utf8").split(b";base64,")[1]
    with open(os.path.join(APP_PATH+"/Data", name), "wb") as fp:
    #with open(os.path.join(APP_PATH, os.path.join("data", "spc_data.csv"))) as fp:
        
        fp.write(base64.decodebytes(data))
    print("Inside Save_file")
    return


#   ------------------------------------Logic for reading uploaded file --------------------------
def uploaded_files():
    """List the files in the upload directory."""
    files = []
    
    for filename in os.listdir(APP_PATH+"/Data"):
        path = os.path.join(APP_PATH+"/Data", filename)
        if os.path.isfile(path):
            #print("upload files")
            files.append(filename)
    return files

#   ------------------------------------Logic for genreating contribution table --------------------------
    
def generate_table_data_new(file_name,condition_flag1):
    
    if condition_flag1:
        
        df_params=pd.read_excel(os.path.join(APP_PATH, os.path.join("data", file_name)),sheet_name="Contribution_table")
        var_names=df_params['Variable']
        df_params=df_params.set_index(df_params.columns[0])
        #print(df_params.columns)
        
        df_params=df_params[['Coeff.','logit_min','logit_average_used','logit_overall_average',
                              'logit_active_average','logit_overall_average_last_year',
                              'logit_active_average_last_year','untransformed_columns']]
    
        
        look_up_index=[2]*len(var_names)
    
        
        
        data_summary_df=pd.read_excel(os.path.join(APP_PATH, os.path.join("data", file_name)),sheet_name="Data Summary")
        kpi_name=pd.read_excel(os.path.join(APP_PATH, os.path.join("data", file_name)),sheet_name="KPI")
        kpi_name=kpi_name['kpi_name'][0]
        
        final_output=contribution_old(df_params,data_summary_df,kpi_name,look_up_index,condition_flag1)
        #print(final_output)
        final_output.insert(0,'Variable',var_names)
        #print(final_output.columns)
        final_output_ob = final_output.to_dict("records")
        #columns=([{'id': p, 'name': p} for p in col_names])
        columns=([{'id': p, 'name': p} for p in final_output.columns])
        final_output.rename(columns={'percentage_normalized':'Contribution'},inplace=True)
        final_output.to_excel(os.path.join(APP_PATH, os.path.join("data", file_name+'_contribution.xlsx')))
        book = load_workbook(os.path.join(APP_PATH, os.path.join("data", file_name)))
        writer = pd.ExcelWriter(os.path.join(APP_PATH, os.path.join("data", file_name)), engine = 'openpyxl')
        writer.book = book
        final_output.to_excel(writer, sheet_name = 'Contribution_table',index=False)
        #df2.to_excel(writer, sheet_name = 'x2')
        writer.save()
        writer.close()
        
        return final_output_ob,columns
        
         
    else:
        
        df=pd.read_excel(os.path.join(APP_PATH, os.path.join("data", file_name)),sheet_name="Params")
        df_params=df.filter(regex="Params_*",axis=1)
        df_params=df_params.head(1)
        df_params.columns=df_params.columns.str.replace("Params_", "")
        df_params=df_params.transpose(copy=True)
        
        
        df=pd.read_excel(os.path.join(APP_PATH, os.path.join("data", file_name)),sheet_name="Params")
        df_temp=df.filter(regex="Params_*",axis=1)
        df=df[df.columns[~df.columns.isin(df_temp.columns)]]
        print("after Params removal",df.columns)
        
        df_temp=df.filter(regex="P_values_*",axis=1)
        df=df[df.columns[~df.columns.isin(df_temp.columns)]]
        
        print("after P values removal",df.columns)
        df=df[df.columns[~df.columns.isin(['index','R_Square','Adj_R_Square'])]]
        variable_transformations=list(df.iloc[0])
        print("after r square removal removal",df.columns)
        print(variable_transformations)
        
        variable_transformations.insert(0,'na')
        print("************VARIABLE TRANSFORMATIONS*****************************")
        print('vt is ',variable_transformations)
        
    #print("DF Params Index",df_params.columns)
    
        print(df_params.head(2))
        print("df_params columns are ",df_params.columns)
        var_names=df_params.index.tolist()
        look_up_index=[2]*len(var_names)
    
        data_summary_df=pd.read_excel(os.path.join(APP_PATH, os.path.join("data", file_name)),sheet_name="Data Summary")
        kpi_name=pd.read_excel(os.path.join(APP_PATH, os.path.join("data", file_name)),sheet_name="KPI")
        kpi_name=kpi_name['kpi_name'][0]
        #print("KPI name is ",kpi_name,"   ",data_summary_df.size)
        
        final_output=contribution_old(df_params,data_summary_df,kpi_name,look_up_index,condition_flag1)
        #print(final_output)
        final_output.insert(0,'Variable',var_names)
        final_output['untransformed_columns']=variable_transformations
        final_output=final_output[["Variable","Coeff.","percentage_normalized","logit_average_used","logit_min",
               "logit_overall_average","logit_active_average","logit_overall_average_last_year",
               "logit_active_average_last_year","predicted_value","predicted_condition","predicted_oa_actual",
               "predicted_condition_actual","percentage_unormalized","untransformed_columns"]]
        
        final_output.rename(columns={'percentage_normalized':'Contribution'},inplace=True)
        
        final_output_ob = final_output.to_dict("records")
        #columns=([{'id': p, 'name': p} for p in col_names])
        columns=([{'id': p, 'name': p} for p in final_output.columns])
        
        final_output.to_excel(os.path.join(APP_PATH, os.path.join("data", file_name+'_contribution.xlsx')),index=False)
        
        book = load_workbook(os.path.join(APP_PATH, os.path.join("data", file_name)))
        writer = pd.ExcelWriter(os.path.join(APP_PATH, os.path.join("data", file_name)), engine = 'openpyxl')
        writer.book = book
        final_output.to_excel(writer, sheet_name = 'Contribution_table')
        #df2.to_excel(writer, sheet_name = 'x2')
        writer.save()
        writer.close()
        
        
        
        return final_output_ob,columns


#--------------------------Logic for calculating contribution ---------------------------------
def contribution_old(par_df, summ_df ,dependent_variable,look_up_index,condition_flag):

    list_con = ['_lag01', '_lag04', '_lag02', '_lag03', '_diff01', '_adstock01', '_adstock1', '_adstock02', '_adstock2', '_adstock03', '_adstock3', '_adstock04', '_adstock4', '_adstock05', '_adstock5', '_adstock06', '_adstock6', '_adstock07', '_adstock7', '_adstock08', '_adstock8', '_ad1', '_ad2', '_ad3', '_ad4', '_ad5', '_ad6', '_ad7','_ad8', '_adstock_1', '_adstock_2',
          '_adstock_3', '_adstock_4', '_adstock_5', '_adstock_6', '_adstock_7', '_adstock_8']

    new_idx = []
    for i in range(len(par_df)):
        k = str(par_df.index[i])
        #print(k)
        for j in list_con:
            if j in k:
                k = k[:(-1*len(j))]
                new_idx.append(k)            
    old_idx = [par_df.index[l] for l in range(len(par_df))]
    for j in range(len(old_idx)):
        for i in new_idx:
            if i in (old_idx[j]):
                old_idx[j] = i
    
    par_df['new_idx'] = old_idx
    par_df.set_index('new_idx', inplace = True)    
    summary = summ_df.set_index(summ_df.columns[0])
    #print("before kvalue")
    kval = summary.loc[str(dependent_variable), 'k_value']
    #print("K-value",kval)
    if condition_flag:
        #par_df['logit_average_used'] = [summary.loc[i,'overall_average'] for i in (par_df.index) ]
        print("function calling via button click")
    else:
        par_df['logit_min'] = [summary.loc[i,'min'] for i in (par_df.index) ]
        par_df['logit_average_used'] = [summary.loc[i,'overall_average'] for i in (par_df.index) ]
        par_df['logit_overall_average'] = [summary.loc[i,'overall_average'] for i in (par_df.index) ]
        par_df['logit_active_average'] = [summary.loc[i,'active_average'] for i in (par_df.index) ]
        par_df['logit_overall_average_last_year'] = [summary.loc[i,'overall_average_last_year'] for i in (par_df.index) ]
        par_df['logit_active_average_last_year'] = [summary.loc[i,'active_average_last_year'] for i in (par_df.index) ]


    custom_index = look_up_index
    
    x=0
    print("len of par df is ",len(par_df))
    for i in range(len(par_df)):
        j = custom_index[i]
        x += par_df.iloc[i,0]*par_df.iloc[i,j]

    
    par_df['predicted_value'] = x
    predicted_oa = par_df['predicted_value']    

    predicted_condition = []
    for i in range(len(par_df)):
        k=custom_index[i]
        val = par_df.iloc[i,0]*par_df.iloc[i,1]
        val1 = par_df.iloc[i,0]*par_df.iloc[i,k]
        val2 = sum(par_df.iloc[:,0]*par_df.iloc[:,2])
        val2 = val2 - val1 + val
        predicted_condition.append(val2)
    val3 = sum(par_df.iloc[:,0]*par_df.iloc[:,1])
    predicted_condition[0] = val3
    
    par_df['predicted_condition'] = predicted_condition
    
    predicted_oa_actual = []
    for i in range(len(par_df)):
        v = np.log(predicted_oa[i]/(1-predicted_oa[i]))*kval
        predicted_oa_actual.append(v)
    
    par_df['predicted_oa_actual']= predicted_oa_actual
    
    predicted_condition_actual = []
    for i in range(len(par_df)):
        v = np.log(predicted_condition[i]/(1-predicted_condition[i]))*kval
        predicted_condition_actual.append(v)
    par_df['predicted_condition_actual']= predicted_condition_actual
        
    percentage_unormalized = []
    for i in range(len(par_df)):
        val = (par_df.iloc[i,9]-par_df.iloc[i,10])/par_df.iloc[i,10]
        percentage_unormalized.append(val)
    par_df['percentage_unormalized'] = percentage_unormalized
    par_df['percentage_normalized'] =percentage_unormalized/np.sum(percentage_unormalized)
    
    if par_df['percentage_normalized'].max() != par_df['percentage_normalized'][0]:
        #print('baseline formula changed')
        par_df.iloc[0,11] = (1-(par_df.iloc[0,9]-par_df.iloc[0,10])/par_df.iloc[0,10])
        percentage_unormalized[0] = (1-(par_df.iloc[0,9]-par_df.iloc[0,10])/par_df.iloc[0,10])
    par_df['percentage_normalized'] =percentage_unormalized/np.sum(percentage_unormalized)
 
    par_df.reset_index(drop=True,inplace=True)
    
    par_df.rename(columns={0:'Coeff.'},inplace=True)
    par_df['percentage_normalized']=par_df['percentage_normalized']*100

    return par_df

#--------------------------call back for contributions adjustment ---------------------------------
@app.callback(
    #[Output('contribution_table', 'data'),Output('contribution_table', 'columns')],
    [Output('contribution_table', 'data'),Output('contribution_table', 'columns'),
 #    Output('contri_table','data'),Output('contri_table','columns')
     ],
    [Input("update_contri_button", "n_clicks"),Input("file_name","children")],
    [State('contribution_table', 'data')]
)
def update_contri_data_table(n_clicks,file_name,contri_table_data):
    print("Inside update_contri_data_table")
    if n_clicks is None or file_name is None:
        if file_name is None:
            print("Default table format")
            return  generate_table_data(),([{'id': p, 'name': p} for p in col_names])
        else:
            print("calling contri func 1st time")
            print(file_name)
            print(type(file_name))
            #print()
            data_ob,columns=generate_table_data_new(file_name['props']['children'],0)
            
            #columns1=([{'id': p, 'name': p} for p in col_names]),
            #data1=generate_table_data(),
            return data_ob,columns
        
        
    else:
        print(file_name)
        print("calling contri func 2nd time")
        #print((contri_table_data))
        
        contri_table_data=pd.DataFrame(contri_table_data)
        #print(contri_table_data)
        #print(type(contri_table_data))
        #print(contri_table_data.columns)
        #columns1=([{'id': p, 'name': p} for p in col_names]),
        #data1=generate_table_data(),
        data_ob,columns=update_contri_table(contri_table_data,file_name['props']['children'],1)
        return data_ob,columns
    
    return [html.Li("xyz")]

def update_contri_table(df_params,file_name,condition_flag1):
    
    
        
    #df_params=pd.read_excel(os.path.join(APP_PATH, os.path.join("data", file_name)),sheet_name="Contribution_table")
    var_names=df_params['Variable']
    transformations_list=df_params['untransformed_columns']
    df_params=df_params.set_index(df_params['Variable'])
    print(df_params.columns)
    print("before")
    df_params=df_params[['Coeff.','logit_min','logit_average_used','logit_overall_average',
                          'logit_active_average','logit_overall_average_last_year','logit_active_average_last_year'
                          ]]

    print(df_params.head(2))
    print("df_params columns are ",df_params.columns)
    df_params=df_params.convert_objects(convert_numeric=True)
    print("dtypes are ",df_params.dtypes)
    look_up_index=[2]*len(var_names)

    
    
    data_summary_df=pd.read_excel(os.path.join(APP_PATH, os.path.join("data", file_name)),sheet_name="Data Summary")
    kpi_name=pd.read_excel(os.path.join(APP_PATH, os.path.join("data", file_name)),sheet_name="KPI")
    kpi_name=kpi_name['kpi_name'][0]
    
    final_contribution_df=contribution_old(df_params,data_summary_df,kpi_name,look_up_index,condition_flag1)
    #print(final_output)
    final_contribution_df.insert(0,'Variable',var_names)
    #print(final_output.columns)
    final_contribution_df=final_contribution_df[["Variable","Coeff.","percentage_normalized","logit_average_used","logit_min",
                   "logit_overall_average","logit_active_average","logit_overall_average_last_year",
                   "logit_active_average_last_year","predicted_value","predicted_condition","predicted_oa_actual",
                   "predicted_condition_actual","percentage_unormalized"]]
    
    final_contribution_df['untransformed_columns']=transformations_list
    final_contribution_df.rename(columns={'percentage_normalized':'Contribution'},inplace=True)
    final_contribution_df.to_excel(os.path.join(APP_PATH, os.path.join("data", file_name+'_contribution.xlsx')),index=False)
    final_contribution_df_ob = final_contribution_df.to_dict("records")
    #columns=([{'id': p, 'name': p} for p in col_names])
    columns=([{'id': p, 'name': p} for p in final_contribution_df.columns])
    
    #adding contribution sheet into the excel file uploaded
    #add_sheet_in_excel_file(file_name,final_contribution_df,"Contribution")
    

    
    return final_contribution_df_ob,columns

#--------------------------Call back for decomposition ---------------------------------
@app.callback(
    #[Output('contribution_table', 'data'),Output('contribution_table', 'columns')],
    [Output('decomp_absolute_table','data'),Output('decomp_absolute_table','columns'),
     Output('decomp_chart','figure'),Output('download_div','children')],
    [Input("update_decomp_chart", "n_clicks")],
    [State("file_name","children"),State('contribution_table', 'data'),]
)
def update_decomp_data(n_clicks,file_name,contri_table_data):
    
    print("Inside update_decomp_data")
    if n_clicks is None or file_name is None:
        if file_name is None:
            print("inside update decomp 1st if")
            return [html.Li("")]
        else:
            print("inside update decomp 1st else")
            print(file_name)
            print(type(file_name))
            #print()
            return [html.Li("")]
        
        
    else:
        
        decomp_absolute_df,decomp_absolute_cols,decomp_percentage_df,decomp_percentage_cols,fig=prepare_and_compute_decomp_data(file_name,contri_table_data)
        download_link=file_download_link(file_name['props']['children'])
        
        #print(raw_data_df.iloc[0])
        #data_ob,columns=update_contri_table(contri_table_data,file_name['props']['children'],1)
        return decomp_percentage_df,decomp_percentage_cols,fig,download_link
    
    
    
    return

def prepare_and_compute_decomp_data(file_name,contri_table_data):
    
    print("file_name",file_name['props']['children'])
    print("calling contri func 2nd time")
    #print((contri_table_data))
    
    #contri_table_data=pd.DataFrame(contri_table_data)
    #print(contri_table_data)
    #print(type(contri_table_data))
    
    #data prep.
    
    contri_table_data=pd.DataFrame(contri_table_data)
    contri_table_data_df=(contri_table_data)
    #print("dtypes are ",contri_table_data.dtypes)
    
    temp_df=pd.DataFrame()
    temp_df['percentage_normalized']=contri_table_data['Contribution']
    temp_df=temp_df.apply(pd.to_numeric, errors='coerce')
    #temp_df['percentage_normalized'].astype(np.float64)
    print("dtypes are ",temp_df.dtypes)
    temp_df['Variable']=contri_table_data['Variable']
    temp_df['transformations']=contri_table_data['untransformed_columns']
    contri_table_data=temp_df
    #print(contri_table_data.columns)
    
    #contri_table_data=contri_table_data[['Variable','percentage_normalized']]
    
    #contri_table_data.convert_objects(convert_numeric=True)
    #contri_table_data['percentage_normalized'].astype(np.float64)
    print("dtypes are ",contri_table_data.dtypes)
    print('columns are',contri_table_data.columns)
    contri_table_data['percentage_normalized']=contri_table_data['percentage_normalized']/100
    print("******************************************************")
    contri_table_data=contri_table_data.iloc[1:]
    
    #reading excel file
    raw_data_df=pd.read_excel(os.path.join(APP_PATH, os.path.join("data", file_name['props']['children'])),sheet_name="Data")
    kpi_name=pd.read_excel(os.path.join(APP_PATH, os.path.join("data", file_name['props']['children'])),sheet_name="KPI")
    kpi_name=kpi_name['kpi_name'][0]
    df_kpi_date=raw_data_df[['Date',kpi_name]]
    
# =============================================================================
#     df_params=pd.read_excel(os.path.join(APP_PATH, os.path.join("data", file_name['props']['children'])),sheet_name="Params")
#     df_temp=df_params.filter(regex="Params_*",axis=1)
#     df_params=df_params[df_params.columns[~df_params.columns.isin(df_temp.columns)]]
#     df_temp=df_params.filter(regex="P_values_*",axis=1)
#     df_params=df_params[df_params.columns[~df_params.columns.isin(df_temp.columns)]]
#     df_params=df_params[df_params.columns[~df_params.columns.isin(['index','R_Square','Adj_R_Square'])]]
#     list_of_variables=list(df_params.iloc[0]) 
#     print()
#     contri_table_data['transformations']=list_of_variables
# =============================================================================
    #print(list_of_variables)
    #list_of_variables.insert(0,kpi_name)
    #list_of_variables.insert(0,'Date')
    
    #print(contri_table_data)
    raw_data_df=raw_data_df[list(contri_table_data['transformations'])]
    print("raw_data_df size is ",raw_data_df.shape)
    #raw_data_df=raw_data_df.insert(0,'Variable',var_names)
    #print(raw_data_df.columns)
    
    #Algo of decomp starts from here 
    sum_of_kpi=sum(df_kpi_date[kpi_name])
    
    mean_of_independent_variables=raw_data_df.mean()
    
    #print(mean_of_independent_variables)
    intermediate_df1=pd.DataFrame()
    intermediate_df1['Yi_by_sum_kpi']=df_kpi_date[kpi_name]/sum_of_kpi
    print("intermediate_df1 size is ",intermediate_df1.shape)
    print(intermediate_df1.columns)
    
    intermediate_df2=pd.DataFrame()
    intermediate_df2=raw_data_df/mean_of_independent_variables
    print("intermediate_df2 size is ",intermediate_df2.shape)
    print(intermediate_df2.columns)
    
    intermediate_df3=pd.DataFrame()
    for col in intermediate_df2.columns:
        print(col)
        #print(intermediate_df2[col]*(intermediate_df1['Yi_by_sum_kpi']))
        intermediate_df3[col]=intermediate_df2[col]*intermediate_df1['Yi_by_sum_kpi']
    #intermediate_df3=intermediate_df2.multiply(intermediate_df1)
    
    print("intermediate_df3 size is ",intermediate_df3.shape)
    
    
    sum_of_intermediate_df3=intermediate_df3.sum(axis=0)
    
    
    
    intermediate_df4=pd.DataFrame()
    intermediate_df4=intermediate_df3/sum_of_intermediate_df3
    print(intermediate_df4.columns)
    print("intermediate_df4 size is ",intermediate_df4.shape)
    
    print(intermediate_df4.head(1))
    
    print("all ok till here")
    
    contri_table_data['absolute_contri_of_each_variable']=contri_table_data['percentage_normalized']*sum_of_kpi
    print(contri_table_data['Variable'])
    
    temp_df=pd.DataFrame()
    decomp_absolute_df=pd.DataFrame()
    for col in intermediate_df4.columns:
        condition_of_matching_columns=contri_table_data['transformations']==col
        
        temp_df=contri_table_data[condition_of_matching_columns]
        print(col)
        print(type(temp_df['absolute_contri_of_each_variable'].iloc[0]))
        #print(intermediate_df4[col]*temp_df['absolute_contri_of_each_variable'].iloc[0])
        decomp_absolute_df[col]=intermediate_df4[col]*temp_df['absolute_contri_of_each_variable'].iloc[0]
        #break
    
    temp_df=pd.DataFrame()
    temp_df['Baseline']=decomp_absolute_df.sum(axis=1)
    temp_df['Baseline']= df_kpi_date[kpi_name]-temp_df['Baseline']
    decomp_absolute_df.insert(0,'Baseline',temp_df['Baseline'])
    
    
    decomp_percentage_df=pd.DataFrame()
    for col in decomp_absolute_df.columns:
        decomp_percentage_df[col]=decomp_absolute_df[col]/df_kpi_date[kpi_name]
        #decomp_percentage_df=decomp_percentage_df*100
    
    
    decomp_absolute_df.insert(0,kpi_name,df_kpi_date[kpi_name])
    decomp_absolute_df.insert(0,'Date',df_kpi_date['Date'])
    
    decomp_percentage_df.insert(0,kpi_name,df_kpi_date[kpi_name])
    decomp_percentage_df.insert(0,'Date',df_kpi_date['Date'])
    
    decomp_absolute_df_ob=decomp_absolute_df.to_dict("records")
    decomp_percentage_df_ob=decomp_percentage_df.to_dict("records")
    #decomp_absolute_df['Date']=raw_data_df['Date']
    #decomp_absolute_df['Date']=raw_data_df[kpi_name]
    #print(decomp_absolute_df.iloc[:,1])
    decomp_absolute_columns=([{'id': p, 'name': p} for p in decomp_absolute_df.columns])
    decomp_percentage_columns=([{'id': p, 'name': p} for p in decomp_percentage_df.columns])
    
# =============================================================================
#     intermediate_df2.to_excel("intermediate_df2.xlsx")
#     intermediate_df3.to_excel("intermediate_df3.xlsx")
#     intermediate_df4.to_excel("intermediate_df4.xlsx")
#     decomp_absolute_df.to_excel("decomp_output.xlsx")
#     decomp_percentage_df.to_excel("decomp_output_percentage.xlsx")
# =============================================================================
    
    fig=update_decomp_chart(decomp_absolute_df,kpi_name)
    print("calculated")
    
    #add_sheet_in_excel_file(file_name['props']['children'],decomp_absolute_df,"Decomp._absolute")
    #add_sheet_in_excel_file(file_name['props']['children'],decomp_percentage_df,"Decomp._percentage")
    add_sheet_in_excel_file(file_name['props']['children'],[contri_table_data_df,decomp_percentage_df,decomp_absolute_df],["Contribution Table","DEcomp. Percentage","DEcomp. Absolute"])
    
    
    
    return decomp_absolute_df_ob,decomp_absolute_columns,decomp_percentage_df_ob,decomp_percentage_columns,fig

#update_decomp_chart
#decomp_div
def update_decomp_chart(decomp_absolute_df,kpi_name):
    
    
    fig =go.Figure()
    for col in decomp_absolute_df.columns:
        if col in['Date','Baseline']:
            print("do nothing")

            
        else:
            if col==kpi_name :
                print('kpi name is ',kpi_name)
                fig.add_trace(go.Line(x=decomp_absolute_df['Date'], y=decomp_absolute_df[col],fill='none', name=col,
                                      mode='lines',line=dict(width=2, color='rgb(0,0,0)'),))
            else:
       
                fig.add_trace(go.Scatter(x=decomp_absolute_df['Date'], y=decomp_absolute_df[col], fill='tozeroy',name=col,stackgroup='one')) # fill down to xaxis
    fig.update_layout(legend_orientation="h")
    return fig

def file_download_link(filename):
    """Create a Plotly Dash 'A' element that downloads a file from the app."""
    location = "/download/{}".format(urlquote(filename))
    
    return html.A("download decomp file here", href=location)

def add_sheet_in_excel_file(file_name,df_list_to_add,sheet_name_list):
    
    book = load_workbook(os.path.join(APP_PATH, os.path.join("data", file_name)))
    writer = pd.ExcelWriter(os.path.join(APP_PATH, os.path.join("data", file_name)), engine = 'openpyxl')
    writer.book = book
    for df_to_add,sheet_name in zip(df_list_to_add,sheet_name_list):
        df_to_add.to_excel(writer, sheet_name = sheet_name)
        #df2.to_excel(writer, sheet_name = 'x2')
    writer.save()
    writer.close()
    return


# Tab 1 callback
@app.callback(
    [Output('datatable-interactivity', 'columns'),Output('datatable-interactivity', 'data'),
     Output('temp_contri_table','data'),Output('temp_contri_table','columns'),
     Output('avp','data'),Output('avp','columns'),
     Output('rsqr','data'),Output('rsqr','columns')],
     
    [Input('tv-column', 'value'),Input('dependent-column', 'value'),Input('var1', 'value'),Input('var2', 'value'),
     Input('var3', 'value'),Input('var4', 'value'),Input('l1', 'value'),Input('l2', 'value'),Input('l3', 'value'),
     Input('l4', 'value'),Input('l5', 'value'),Input('base', 'value'), Input('var5', 'value'),Input('var6', 'value'),
     Input('var7', 'value'),Input('var8', 'value'),Input('var9', 'value'),Input('var10', 'value'),Input('l6', 'value'),
     Input('l7', 'value'), Input('l8', 'value'), Input('l9', 'value'), Input('l10', 'value'), Input('l11', 'value')],
     [State('summary_table','data'),State('dfl_table','data'),State('dfa_table','data')])
def regression(tv_var, depen, var1, var2, var3, var4, l1, l2, l3, l4, l5, base, var5, var6, var7, var8, var9, var10, l6, l7, l8, l9, l10, l11, summary_dict, dfl_dict,dfa_dict):
    
         logit_index = [2, int(l1), int(l2), int(l3), int(l4), int(l5), int(l6), int(l7), int(l8), int(l9), int(l10), int(l11)]
         string_list = [str(tv_var), str(var1), str(var2), str(var3), str(var4), str(var5), str(var6), str(var7), str(var8), str(var9), str(var10)]
         string_list = [i for i in string_list if i != 'null']
         form1 = " + ".join(string_list)
         formula = str(depen)+ ' ' + '~' + form1  
    
         dfl=pd.DataFrame(dfl_dict)
         print("all ok till here ***********************************************************")
         print(dfl.head(1))
         print("Summary ***********************************************************")
    
         results = sm.ols(formula, data=dfl).fit()
         df_p = pd.DataFrame(results.pvalues)
         df_e = pd.DataFrame(results.params)
    
         summary=pd.DataFrame(summary_dict)
         summary.set_index('var_names', inplace = True)
         del summary.index.name
         print(summary.head(1))
         par_dff, par_df = contribution(df_e, summary , str(depen), df_p, logit_index, base)
         resu = pd.DataFrame(par_dff)
         columns=[{"name": i, "id": i, "deletable": True, "selectable": True} for i in resu.columns]
         data=resu.to_dict('records')
         print(par_df, 'vangaaa')
         resu1=pd.DataFrame(par_df)
         print(resu1.columns)
         
         #Preparing R-Square & AvP
         rsquare = results.rsquared
         avp = pd.DataFrame(results.fittedvalues, columns = ['predicted'])
         avp['actual'] = dfl[str(depen)]
         rsqr = pd.DataFrame([rsquare], columns = ['rsquare'])
         rsqr['k_value'] = summary.loc[str(depen), 'k_value']
         dfa = pd.DataFrame(dfa_dict)
         avp['date'] = dfa['date']
         k = float(rsqr['k_value'])
         avp['predicted'] =  np.log(avp['predicted'].values/(1-avp['predicted'].values))*k
         avp['actual'] = np.log(avp['actual'].values/(1-avp['actual'].values))*k
         avp['date'] = [i[0:-9] for i in avp['date']]
        
         avp_col = [{"name": i, "id": i, "deletable": True, "selectable": True} for i in avp.columns]
         avp_dat = avp.to_dict('records') 
                                
         rsqr_col = [{"name": i, "id": i, "deletable": True, "selectable": True} for i in rsqr.columns]
         rsqr_dat = rsqr.to_dict('records') 
         
         data_to_temp_contri_table=resu1.to_dict('records')
         columns_to_temp_contri_table=[{"name": i, "id": i, "deletable": True, "selectable": True} for i in resu1.columns]
         
         
         
         return columns, data ,data_to_temp_contri_table,columns_to_temp_contri_table,avp_dat,avp_col,rsqr_dat,rsqr_col

#===========================AvP & R-square callback=============================
@app.callback(
    dash.dependencies.Output('avp_graph', 'figure'),    
    #[Input('show_avp', 'n_clicks'),],
    [Input('avp', 'data')],[State('rsqr', 'data')]
    
    )

def avp_plot(avp, rsqr ):
        # Create figure with secondary y-axis
        print("Inside AVP callback")
        print(rsqr)
        avp = pd.DataFrame(avp)
        rsqr = pd.DataFrame(rsqr)
        avp.to_csv('avpcheck.csv')
        rsqr.to_csv('rsqrcheck.csv') 
        fig = go.Figure()
        fig.add_trace(
                go.Scatter(
                        x=avp['date'],
                        y=avp['predicted'],
                        name = 'Predicted KPI'
                        ))
        fig.add_trace(
                go.Scatter(
                        x=avp['date'],
                        y=avp['actual'],
                        name = 'Actual KPI'
                        ))
        
        fig.update_layout(
                title= 'R Square of Model = '+ str(rsqr['rsquare'][0]*100)[0:5],
                xaxis_title="Actual versus Predicted")

        return(fig)    
   
    
#===============callback to freeze contributions of tab1 =================================
@app.callback([Output('contri_table','data'),Output('contri_table','columns')],[Input('freeze_contribution','n_clicks')],
               [State('temp_contri_table','data'),State("file_name","children")])
def update_contri_table_from_tab1(n_clicks,contribution_dict,file_name):
    contribution_df=pd.DataFrame(contribution_dict)
    print("All OK **************")
    contribution_df=contribution_df[["Variable","Coeff.","percentage_normalized","logit_average_used","logit_min",
                   "logit_overall_average","logit_active_average","logit_overall_average_last_year",
                   "logit_active_average_last_year","predicted_value","predicted_condition","predicted_oa_actual",
                   "predicted_condition_actual","percentage_unormalized","untransformed_columns"]]
    contribution_df.rename(columns={'percentage_normalized':'Contribution'},inplace=True)    
    contribution_df.to_excel(os.path.join(APP_PATH, os.path.join("data", file_name['props']['children']+'_contribution.xlsx')),index=False)
    contribution_dict=contribution_df.to_dict("records")
    columns=[{"name": i, "id": i} for i in contribution_df.columns ]
    return contribution_dict,columns


@app.callback(
    [Output('correlation_matrix', 'columns'),
     Output('correlation_matrix', 'data')],
    [Input('tv-column', 'value'),
     Input('dependent-column', 'value'),
     Input('var1', 'value'),
     Input('var2', 'value'),
     Input('var3', 'value'),
     Input('var4', 'value'),
     ],[State('dfl_table','data')])

def matrix(tv_var, depen, var1, var2, var3, var4,dfl_dict):
    
    string_list = [str(tv_var), str(var1), str(var2), str(var3), str(var4)]
    string_list = [i for i in string_list if i != 'null']
    dfl=pd.DataFrame(dfl_dict)
    data_for_matrix = dfl[string_list]  
    matrix = data_for_matrix.corr()
    matrix['variable'] = [i for i in matrix.index]
    cols = list(matrix.columns)
    cols = [cols[-1]] + cols[:-1]
    matrix = matrix[cols]
    columns=[
            {"name": i, "id": i, "deletable": True, "selectable": True} for i in matrix.columns
        ]
    data=matrix.to_dict('records')    
    return columns, data 




#====================== Tab 2 callback
   
    
@app.callback(
    dash.dependencies.Output('crossfilter-indicator-scatter', 'figure'),    
    [Input('plot_dependent', 'value'),Input('plot_independent', 'value')],[State('dfa_table','data')]
    )

def plot(depen,indepen,dfa_dict):
    # Create figure with secondary y-axis
    dfa=pd.DataFrame(dfa_dict)
    #fig1 = go.Figure()
    fig1 = make_subplots(specs=[[{"secondary_y": True}]])
    
    # Add traces
    
    fig1.add_trace(go.Scatter(x= dfa['date'], y=dfa[str(depen)], name=depen),secondary_y=False,)
    fig1.add_trace(go.Scatter(x= dfa['date'], y=dfa[str(indepen)], name=indepen),secondary_y=True,)
    # Margins
    fig1.update_layout(legend_orientation="h")

    # Set x-axis title
    fig1.update_xaxes(title_text="Date", tickfont = dict(size = 10))
    
    # Set y-axes titles
    fig1.update_yaxes(title_text="Y axis", secondary_y=True)
    
    
   
    return(fig1)


@app.callback(
    [Output('datatable-interactivity2', 'columns'),
     Output('datatable-interactivity2', 'data')],    
    [Input('plot_dependent', 'value')],[State('dfa_table','data')]
    )

def cormatrix(depen,dfa_dict):
    dfa=pd.DataFrame(dfa_dict)
    mat = dfa.corr()
    mat['var'] = [i for i in mat.index]
    mat = mat[['var', str(depen)]]
    mat = mat.sort_values([str(depen)], ascending = False)
#    mat = pd.DataFrame(mat[str(depen)]
    columns=[
            {"name": i, "id": i, "deletable": True, "selectable": True} for i in mat.columns
        ]
    data=mat.to_dict('records')     
    return columns, data 


# Running the server
if __name__ == "__main__":
    app.run_server(debug=False, port=8050)
